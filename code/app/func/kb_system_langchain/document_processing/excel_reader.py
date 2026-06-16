"""Excel 文件读取工具"""

from pathlib import Path
from typing import Dict, List, Optional


def load_xlsx_rows(file_path: str) -> List[Dict[str, str]]:
    """
    加载 Excel，第一行为表头，后续每行转为字典。
    .xlsx 使用 openpyxl，.xls 使用 xlrd。
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".xls":
        try:
            import xlrd
        except ImportError:
            raise ImportError("请安装 xlrd: pip install xlrd")

        wb = xlrd.open_workbook(file_path)
        result: List[Dict[str, str]] = []
        for sheet in wb.sheets():
            if sheet.nrows < 2:
                continue
            headers = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            headers = [h if h.strip() else f"col_{i}" for i, h in enumerate(headers)]
            sheet_name = sheet.name
            for row_idx in range(1, sheet.nrows):
                row_dict = {"_sheet": sheet_name} if wb.nsheets > 1 else {}
                for col_idx in range(sheet.ncols):
                    if sheet.cell_type(row_idx, col_idx) != xlrd.XL_CELL_EMPTY:
                        val = sheet.cell_value(row_idx, col_idx)
                        row_dict[headers[col_idx]] = str(val)
                if any(v.strip() for v in row_dict.values() if v and v != sheet_name):
                    result.append(row_dict)
        return result

    try:
        import openpyxl
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    result: List[Dict[str, str]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue
        headers = [str(v) if v is not None else "" for v in all_rows[0]]
        headers = [h.strip() if h.strip() else f"col_{i}" for i, h in enumerate(headers)]
        for row in all_rows[1:]:
            row_dict = {"_sheet": sheet_name} if len(wb.sheetnames) > 1 else {}
            for i, val in enumerate(row):
                if val is not None and str(val).strip():
                    row_dict[headers[i]] = str(val)
            if any(v.strip() for k, v in row_dict.items() if k != "_sheet"):
                result.append(row_dict)
    wb.close()
    return result


def resolve_qa_column(
    headers: List[str],
    aliases: tuple,
    explicit: Optional[str] = None,
) -> Optional[str]:
    if explicit:
        explicit = explicit.strip()
        if explicit in headers:
            return explicit
        raise ValueError(f"指定的列 '{explicit}' 不存在，可用列: {', '.join(headers)}")

    header_map = {h.strip(): h for h in headers if h and h != "_sheet"}
    for alias in aliases:
        if alias in header_map:
            return header_map[alias]
    return None


def load_qa_pairs(
    file_path: str,
    q_column: Optional[str] = None,
    a_column: Optional[str] = None,
) -> List[tuple[str, str]]:
    from func.kb_system_langchain.document_processing.constants import (
        QA_A_COLUMN_ALIASES,
        QA_Q_COLUMN_ALIASES,
    )

    row_dicts = load_xlsx_rows(file_path)
    if not row_dicts:
        return []

    sample_keys = [k for k in row_dicts[0].keys() if k != "_sheet"]
    q_col = resolve_qa_column(sample_keys, QA_Q_COLUMN_ALIASES, q_column)
    a_col = resolve_qa_column(sample_keys, QA_A_COLUMN_ALIASES, a_column)

    if not q_col:
        raise ValueError(
            f"未找到 Q 列，请确保表头含 {QA_Q_COLUMN_ALIASES} 之一，或通过 q_column 指定"
        )
    if not a_col:
        raise ValueError(
            f"未找到 A 列，请确保表头含 {QA_A_COLUMN_ALIASES} 之一，或通过 a_column 指定"
        )

    pairs: List[tuple[str, str]] = []
    for row in row_dicts:
        question = str(row.get(q_col, "")).strip()
        answer = str(row.get(a_col, "")).strip()
        if question:
            pairs.append((question, answer))
    return pairs
